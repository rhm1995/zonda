"""TASK-001 unit tests: workbook parsing against the real bundled files
(spot-check regression) and a deliberately corrupted fixture (fail-fast
path, acceptance criterion 4)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from data_pipeline.parse_ons_workbook import WorkbookStructureError, parse_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
NEWBUILD_PATH = REPO_ROOT / "data" / "raw" / "newbuild.xlsx"
EXISTING_PATH = REPO_ROOT / "data" / "raw" / "existing.xlsx"


def test_parses_real_newbuild_workbook_with_confirmed_spot_checks() -> None:
    parsed = parse_workbook(NEWBUILD_PATH, "new_build")
    assert len(parsed.local_authorities) == 318
    assert len(parsed.period_labels) == 120
    assert len(parsed.price_points) == 318 * 120

    def price(la_name: str, period_label: str) -> int | None:
        matches = [
            p
            for p in parsed.price_points
            if p.la_name == la_name and p.period_label == period_label
        ]
        assert len(matches) == 1
        return matches[0].price_gbp

    assert price("Manchester", "Year ending Sep 2025") == 495000
    assert price("Manchester", "Year ending Sep 2015") == 177995


def test_parses_real_existing_workbook_with_confirmed_spot_checks() -> None:
    parsed = parse_workbook(EXISTING_PATH, "existing")

    def price(la_name: str, period_label: str) -> int | None:
        matches = [
            p
            for p in parsed.price_points
            if p.la_name == la_name and p.period_label == period_label
        ]
        assert len(matches) == 1
        return matches[0].price_gbp

    assert price("Manchester", "Year ending Sep 2025") == 400000
    assert price("Manchester", "Year ending Sep 2015") == 220000


def test_geography_is_england_and_wales_only_no_scotland_or_ni() -> None:
    parsed = parse_workbook(NEWBUILD_PATH, "new_build")
    region_codes = {la.region_country_code for la in parsed.local_authorities}
    assert all(code.startswith("E12") or code.startswith("W92") for code in region_codes)


def test_suppressed_marker_is_recognised_and_carries_no_price() -> None:
    parsed = parse_workbook(NEWBUILD_PATH, "new_build")
    suppressed_points = [p for p in parsed.price_points if p.suppressed]
    assert suppressed_points, "expected at least one suppressed cell in the real workbook"
    assert all(p.price_gbp is None for p in suppressed_points)


def _write_corrupted_workbook(path: Path, *, wrong_column_count: bool) -> None:
    """A minimal fixture workbook that diverges from the confirmed layout
    (design §6.1): either a malformed period-label header or a wrong period
    column count, simulating ONS-side layout drift."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2b"
    ws.append(["Table 2b - fixture"])
    ws.append(["Source: fixture"])
    headers = ["Region/Country code", "Region/Country name", "Local authority code", "Local authority name"]
    if wrong_column_count:
        headers += ["Year ending Sep 2024", "Year ending Sep 2025"]  # only 2, not 120
    else:
        headers += ["Not a period label", "Year ending Sep 2025"]
    ws.append(headers)
    ws.append(["E12000001", "North East", "E06000001", "Hartlepool", 100000, 105000])
    wb.save(path)


def test_wrong_period_column_count_fails_the_build_with_a_specific_diagnostic(tmp_path: Path) -> None:
    fixture_path = tmp_path / "corrupted.xlsx"
    _write_corrupted_workbook(fixture_path, wrong_column_count=True)
    with pytest.raises(WorkbookStructureError, match="120 period columns"):
        parse_workbook(fixture_path, "new_build")


def test_malformed_period_label_fails_the_build_with_a_specific_diagnostic(tmp_path: Path) -> None:
    fixture_path = tmp_path / "corrupted.xlsx"
    _write_corrupted_workbook(fixture_path, wrong_column_count=False)
    with pytest.raises(WorkbookStructureError, match="Unrecognised period label|120 period columns"):
        parse_workbook(fixture_path, "new_build")


def test_missing_sheet_fails_loudly(tmp_path: Path) -> None:
    fixture_path = tmp_path / "no_2b.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "not_2b"
    wb.save(fixture_path)
    with pytest.raises(WorkbookStructureError, match="2b"):
        parse_workbook(fixture_path, "new_build")


def test_unexpected_cell_value_fails_loudly(tmp_path: Path) -> None:
    fixture_path = tmp_path / "bad_cell.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2b"
    ws.append(["Table 2b - fixture"])
    ws.append(["Source: fixture"])
    headers = ["Region/Country code", "Region/Country name", "Local authority code", "Local authority name"]
    headers += [f"Year ending {label}" for label in _fixture_period_labels()]
    ws.append(headers)
    row = ["E12000001", "North East", "E06000001", "Hartlepool"] + ["N/A"] * 120
    ws.append(row)
    wb.save(fixture_path)
    with pytest.raises(WorkbookStructureError, match="non-numeric"):
        parse_workbook(fixture_path, "new_build")


def _fixture_period_labels() -> list[str]:
    """120 valid, chronologically ordered period labels for a synthetic
    fixture workbook (matching the real files' Dec1995-Sep2025 axis)."""
    labels = []
    year, months = 1995, ["Dec"]
    labels.append(f"{months[0]} {year}")
    year = 1996
    while len(labels) < 120:
        for month in ["Mar", "Jun", "Sep", "Dec"]:
            labels.append(f"{month} {year}")
            if len(labels) == 120:
                break
        year += 1
    return labels

"""Parse tab 2b of an ONS "Median house prices for administrative
geographies" workbook (newly built or existing dwellings, detached houses)
into long-format `PricePoint` rows (TASK-001, design §6.4).

Offline, build-time only — this module (and `pandas`/`openpyxl` generally)
is never imported by the running app (CON-008). Runtime reads go through
`core/repository.py`'s DuckDB views over the Parquet output this module
produces.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Literal

import openpyxl

from core.models import Dataset, LocalAuthority, PricePoint

SHEET_NAME = "2b"
HEADER_ROW = 3
FIRST_DATA_ROW = 4
FIRST_PERIOD_COLUMN = 5  # columns A-D are region/LA code/name; E.. are periods
EXPECTED_PERIOD_COUNT = 120
SUPPRESSED_MARKER = "[x]"

# "Year ending <Mon> <YYYY>" -> the period's end date, per design §6.4's
# fixed quarterly rolling-year-ending convention.
_MONTH_TO_END_DATE = {
    "Mar": (3, 31),
    "Jun": (6, 30),
    "Sep": (9, 30),
    "Dec": (12, 31),
}
_PERIOD_LABEL_RE = re.compile(r"^Year ending (Mar|Jun|Sep|Dec) (\d{4})$")

# Curated common alternate spellings for local authorities whose ONS-style
# name (e.g. "Kingston upon Hull, City of") differs materially from how a
# question is likely to name it. Not used by Increment 1 (the deterministic
# tabs use closed selectors, ADR-012) — built now for TASK-011's later use,
# per this task's stated scope.
CURATED_ALIASES: dict[str, list[str]] = {
    "Kingston upon Hull, City of": ["Hull", "Kingston upon Hull"],
    "Herefordshire, County of": ["Herefordshire"],
    "Bristol, City of": ["Bristol"],
    "City of London": ["London"],
}


class WorkbookStructureError(Exception):
    """Raised when a workbook's structure diverges from the confirmed
    layout (design §6.1) in a way that would silently corrupt parsed
    figures if ignored — never a silent partial parse (RSK-003)."""


@dataclass(frozen=True)
class ParsedWorkbook:
    dataset: Dataset
    price_points: list[PricePoint]
    local_authorities: list[LocalAuthority]
    period_labels: list[str]  # ordered as they appear in the workbook


def _period_end_date(label: str) -> date:
    match = _PERIOD_LABEL_RE.match(label)
    if not match:
        raise WorkbookStructureError(
            f"Unrecognised period label {label!r}; expected "
            f"'Year ending <Mar|Jun|Sep|Dec> <YYYY>'"
        )
    month_abbr, year_str = match.groups()
    month, day = _MONTH_TO_END_DATE[month_abbr]
    return date(int(year_str), month, day)


def _read_period_headers(header_row: tuple) -> list[str]:
    raw = list(header_row[FIRST_PERIOD_COLUMN - 1 :])
    # A known ONS export artifact: a trailing blank column with no header.
    # Tolerated only at the end of the row — a blank header anywhere else
    # is a structural problem, not this known quirk, and must fail loudly.
    while raw and raw[-1] is None:
        raw.pop()
    if any(header is None for header in raw):
        raise WorkbookStructureError(
            "Found a blank period-label header among data columns "
            "(not just a trailing one) — workbook layout has changed"
        )
    if len(raw) != EXPECTED_PERIOD_COUNT:
        raise WorkbookStructureError(
            f"Expected exactly {EXPECTED_PERIOD_COUNT} period columns, "
            f"found {len(raw)}"
        )
    for header in raw:
        _period_end_date(header)  # raises WorkbookStructureError if malformed
    return raw


def _parse_cell_value(raw_value: object, *, la_name: str, period_label: str) -> tuple[int | None, bool]:
    """Returns (price_gbp, suppressed). Fails loudly on any value that is
    neither a whole-number price nor the confirmed suppression marker
    (design §6.4 — a new/unexpected suppression convention must not be
    silently absorbed as a numeric zero or skipped)."""
    if raw_value == SUPPRESSED_MARKER:
        return None, True
    if isinstance(raw_value, bool):
        raise WorkbookStructureError(
            f"Unexpected boolean cell for {la_name!r} at {period_label!r}"
        )
    if isinstance(raw_value, int):
        return raw_value, False
    if isinstance(raw_value, float) and raw_value.is_integer():
        return int(raw_value), False
    raise WorkbookStructureError(
        f"Unexpected non-numeric, non-{SUPPRESSED_MARKER!r} cell value "
        f"{raw_value!r} for {la_name!r} at {period_label!r}"
    )


def parse_workbook(path: Path, dataset: Dataset) -> ParsedWorkbook:
    """Reshape one workbook's tab 2b from wide to long format.

    Uses `iter_rows(values_only=True)` for a single forward sequential scan
    -- openpyxl's read-only mode is optimised for this access pattern, not
    for per-cell random access via `sheet.cell(row=, column=)`, which is
    dramatically slower at this sheet's size (~76k data cells)."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise WorkbookStructureError(
            f"{path}: expected a sheet named {SHEET_NAME!r}, found {workbook.sheetnames}"
        )
    sheet = workbook[SHEET_NAME]
    row_iter = sheet.iter_rows(values_only=True)
    for _ in range(HEADER_ROW - 1):
        next(row_iter)  # skip title/source rows
    header_row = next(row_iter)
    period_labels = _read_period_headers(header_row)
    period_end_dates = [_period_end_date(label) for label in period_labels]
    last_period_col_index = FIRST_PERIOD_COLUMN - 1 + len(period_labels)

    price_points: list[PricePoint] = []
    local_authorities: dict[str, LocalAuthority] = {}
    for row_number, row in enumerate(row_iter, start=FIRST_DATA_ROW):
        region_code, region_name, la_code, la_name = row[:4]
        if la_code is None and la_name is None:
            break  # end of data
        if not all([region_code, region_name, la_code, la_name]):
            raise WorkbookStructureError(
                f"{path}: row {row_number} has a partially-populated identity "
                f"column set — refusing to guess: "
                f"{(region_code, region_name, la_code, la_name)!r}"
            )
        local_authorities[la_code] = LocalAuthority(
            la_code=la_code,
            la_name=la_name,
            region_country_code=region_code,
            region_country_name=region_name,
            aliases=CURATED_ALIASES.get(la_name, []),
        )
        row_values = row[FIRST_PERIOD_COLUMN - 1 : last_period_col_index]
        for period_label, period_end_date, raw_value in zip(
            period_labels, period_end_dates, row_values
        ):
            price_gbp, suppressed = _parse_cell_value(
                raw_value, la_name=la_name, period_label=period_label
            )
            price_points.append(
                PricePoint(
                    dataset=dataset,
                    region_country_code=region_code,
                    region_country_name=region_name,
                    la_code=la_code,
                    la_name=la_name,
                    period_label=period_label,
                    period_end_date=period_end_date,
                    price_gbp=price_gbp,
                    suppressed=suppressed,
                )
            )

    return ParsedWorkbook(
        dataset=dataset,
        price_points=price_points,
        local_authorities=list(local_authorities.values()),
        period_labels=period_labels,
    )
